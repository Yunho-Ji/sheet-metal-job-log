import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def export_to_excel(logs, file_path, work_date_str, worker_name_str):
    """
    주어진 작업 내역 목록(logs)을 명세서 규격에 맞게 A4 크기 엑셀 파일로 생성합니다.
    
    :param logs: DB에서 가져온 작업일지 딕셔너리 리스트
    :param file_path: 저장할 엑셀 파일 경로
    :param work_date_str: 작성일 문자열 (예: '2026년 07월 27일 월요일')
    :param worker_name_str: 작업자 성명 문자열 (예: '홍길동')
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "절곡작업일지"
    
    # 그리드라인 보이기 설정
    ws.views.sheetView[0].showGridLines = True
    
    # 폰트 설정 정의
    font_title = Font(name="맑은 고딕", size=18, bold=True)
    font_header_bold = Font(name="맑은 고딕", size=10, bold=True)
    font_data = Font(name="맑은 고딕", size=10)
    
    # 테두리 설정 정의
    thin_side = Side(style='thin', color='000000')
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # 배경색 정의 (연한 회색 헤더용)
    fill_header = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # 정렬 정의
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center')
    
    # --- 1. Row 1 ~ 5 상단 헤더 영역 ---
    # A1:D5 병합 - 문서 제목
    ws.merge_cells("A1:D5")
    title_cell = ws["A1"]
    title_cell.value = "생산1팀       절곡       작업 일지"
    title_cell.font = font_title
    title_cell.alignment = align_center
    
    # E1:E5 병합 - 결재 세로 헤더
    ws.merge_cells("E1:E5")
    approval_header = ws["E1"]
    approval_header.value = "결\n\n재"
    approval_header.font = font_header_bold
    approval_header.alignment = align_center
    
    # F1~H1 결재자 직급 헤더
    headers_approval = [("F1", "작성"), ("G1", "검토"), ("H1", "승인")]
    for cell_idx, text in headers_approval:
        cell = ws[cell_idx]
        cell.value = text
        cell.font = font_header_bold
        cell.alignment = align_center
        
    # F2:F5, G2:G5, H2:H5 세로 병합 (서명란)
    ws.merge_cells("F2:F5")
    ws.merge_cells("G2:G5")
    ws.merge_cells("H2:H5")
    
    # 결재 영역 전체에 테두리 적용 (병합된 셀들도 테두리가 정상 적용되도록 전체 셀 순회)
    for r in range(1, 6):
        for c in range(1, 9):  # A~H (1~8)
            ws.cell(row=r, column=c).border = border_all
            
    # --- 2. Row 6 ~ 7 기본 작성 정보 ---
    # Row 6: 작성일
    ws.merge_cells("A6:B6")
    date_header = ws["A6"]
    date_header.value = "작성일"
    date_header.font = font_header_bold
    date_header.alignment = align_center
    
    ws.merge_cells("C6:H6")
    date_val = ws["C6"]
    date_val.value = work_date_str
    date_val.font = font_header_bold
    date_val.alignment = align_center
    
    # 테두리 전체 적용
    for c in range(1, 9):
        ws.cell(row=6, column=c).border = border_all
        
    # Row 7: 작업인원
    ws.merge_cells("A7:B7")
    worker_header = ws["A7"]
    worker_header.value = "작업인원"
    worker_header.font = font_header_bold
    worker_header.alignment = align_center
    
    ws.merge_cells("C7:H7")
    worker_val = ws["C7"]
    worker_val.value = worker_name_str
    worker_val.font = font_header_bold
    worker_val.alignment = align_center
    
    # 테두리 전체 적용
    for c in range(1, 9):
        ws.cell(row=7, column=c).border = border_all
        
    # --- 3. Row 8 ~ 9 메인 작업 데이터 테이블 헤더 ---
    # 헤더 정의 및 셀 병합
    ws.merge_cells("A8:C8")
    ws["A8"] = "작업시간"
    
    ws["A9"] = "시작"
    ws["B9"] = "완료"
    ws["C9"] = "시간"
    
    # 2행씩 세로 병합하는 헤더들
    vertical_headers = [
        ("D8", "D9", "로트번호 및 구분"),
        ("E8", "E9", "제품명"),
        ("F8", "F9", "공정 및 도변"),
        ("G8", "G9", "작업수량"),
        ("H8", "H9", "비고")
    ]
    
    for top_idx, bot_idx, text in vertical_headers:
        ws.merge_cells(f"{top_idx}:{bot_idx}")
        ws[top_idx] = text
        
    # 테이블 헤더 스타일 적용 (Row 8, Row 9)
    for r in [8, 9]:
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.font = font_header_bold
            cell.alignment = align_center
            cell.border = border_all

    # --- 4. Row 10 ~ 데이터 바인딩 ---
    current_row = 10
    for log in logs:
        # 데이터 바인딩
        ws.cell(row=current_row, column=1, value=log.get("start_time", ""))
        ws.cell(row=current_row, column=2, value=log.get("end_time", ""))
        ws.cell(row=current_row, column=3, value=log.get("duration_time", ""))
        ws.cell(row=current_row, column=4, value=log.get("lot_number", ""))
        ws.cell(row=current_row, column=5, value=log.get("product_name", ""))
        ws.cell(row=current_row, column=6, value=log.get("process_code", ""))
        
        # 수량은 숫자로 형변환하여 입력 (천단위 포맷 적용 위해)
        qty = log.get("quantity", "")
        try:
            qty_val = int(qty)
            qty_cell = ws.cell(row=current_row, column=7, value=qty_val)
            qty_cell.number_format = '#,##0'
        except (ValueError, TypeError):
            qty_cell = ws.cell(row=current_row, column=7, value=str(qty))
        
        ws.cell(row=current_row, column=8, value=log.get("remarks", ""))
        
        # 데이터 셀 테두리 및 정렬 스타일 적용
        ws.row_dimensions[current_row].height = 22
        for c in range(1, 9):
            cell = ws.cell(row=current_row, column=c)
            cell.font = font_data
            cell.border = border_all
            
            # 열별 맞춤 정렬
            if c in [1, 2, 3]:    # 시작, 완료, 소요
                cell.alignment = align_center
            elif c == 4:          # 로트번호
                cell.alignment = align_center
            elif c == 5:          # 제품명
                cell.alignment = align_left
            elif c == 6:          # 공정 및 도번
                cell.alignment = align_center
            elif c == 7:          # 수량
                cell.alignment = align_right
            elif c == 8:          # 비고
                cell.alignment = align_left
                
        current_row += 1

    # --- 5. A4 인쇄 설정 및 열 너비 최적화 ---
    # 상단 및 헤더 행 높이 설정
    for r in range(1, 6):
        ws.row_dimensions[r].height = 20
    ws.row_dimensions[6].height = 25
    ws.row_dimensions[7].height = 25
    ws.row_dimensions[8].height = 22
    ws.row_dimensions[9].height = 22
    
    # 기본 열 너비 설정 (한글 텍스트 고려 넉넉하게 지정)
    column_widths = {
        'A': 12,  # 시작시간
        'B': 12,  # 완료시간
        'C': 12,  # 소요시간
        'D': 18,  # 로트번호
        'E': 25,  # 제품명
        'F': 20,  # 공정 및 도번
        'G': 12,  # 수량
        'H': 20   # 비고
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # 인쇄 레이아웃 세부 설정 (A4 세로 맞춤)
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    
    # 여백 설정 (인치 단위: 0.5인치는 약 1.27cm)
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3
    
    # 1페이지 너비에 맞추기 설정
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # 세로 길이는 페이지 내용에 따라 자동으로 넘어감
    
    # 저장
    wb.save(file_path)
    wb.close()
